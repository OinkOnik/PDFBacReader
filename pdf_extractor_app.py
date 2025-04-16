# pdf_extractor_app.py
"""
Clase principal de la interfaz gráfica mejorada para la aplicación de extracción de datos PDF.
Presenta una interfaz más intuitiva y profesional con mejor organización de elementos.

Módulos relacionados:
- pdf_processor.py: Contiene la clase para procesar PDFs en segundo plano
"""

import os
from PyQt6.QtWidgets import (QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout,
                             QFileDialog, QLabel, QTableWidget, QTableWidgetItem,
                             QWidget, QProgressBar, QMessageBox, QGroupBox,
                             QSplitter, QFrame, QStatusBar, QHeaderView)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QIcon, QFont, QAction
from pdf_processor import PDFExtractorThread
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


class PDFExtractorApp(QMainWindow):
    """Aplicación principal mejorada para extraer datos de PDFs"""

    def __init__(self):
        """Inicializa la ventana principal y los componentes de la interfaz"""
        super().__init__()
        self.setWindowTitle("Extractor de Datos PDF - Sistema de Gestión")
        self.setGeometry(100, 100, 1000, 800)  # Ventana más grande
        self.pdf_files = []
        self.original_df = None

        # Crear barra de estado
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("Listo para iniciar")

        # Configurar interfaz
        self.init_ui()

    def init_ui(self):
        """Configura los elementos de la interfaz de usuario mejorada"""
        # Widget central y layout principal
        central_widget = QWidget()
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Crear un splitter para separar los controles de la tabla de resultados
        splitter = QSplitter(Qt.Orientation.Vertical)

        # =============== ÁREA DE CONTROLES ===============
        controls_container = QWidget()
        controls_layout = QVBoxLayout(controls_container)
        controls_layout.setSpacing(15)

        # Sección de archivos
        file_group = QGroupBox("Selección de archivos")
        file_group.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        file_layout = QVBoxLayout(file_group)

        # Área de información de archivos
        file_info_layout = QHBoxLayout()
        self.file_label = QLabel("No hay archivos seleccionados")
        self.file_label.setFont(QFont("Arial", 9))
        self.file_count_label = QLabel("")
        self.file_count_label.setFont(QFont("Arial", 9, QFont.Weight.Bold))
        file_info_layout.addWidget(self.file_label)
        file_info_layout.addWidget(self.file_count_label)
        file_info_layout.addStretch()
        file_layout.addLayout(file_info_layout)

        # Botones de acción para archivos
        file_btn_layout = QHBoxLayout()
        self.select_btn = QPushButton("Seleccionar PDFs")
        self.select_btn.setMinimumHeight(30)
        self.select_btn.clicked.connect(self.select_pdfs)

        self.clear_btn = QPushButton("Limpiar selección")
        self.clear_btn.setMinimumHeight(30)
        self.clear_btn.clicked.connect(self.clear_selection)
        self.clear_btn.setEnabled(False)

        file_btn_layout.addWidget(self.select_btn)
        file_btn_layout.addWidget(self.clear_btn)
        file_layout.addLayout(file_btn_layout)

        controls_layout.addWidget(file_group)

        # Sección de procesamiento
        process_group = QGroupBox("Procesamiento")
        process_group.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        process_layout = QVBoxLayout(process_group)

        # Botón de procesamiento
        self.process_btn = QPushButton("Procesar PDFs")
        self.process_btn.setMinimumHeight(40)
        self.process_btn.setEnabled(False)
        self.process_btn.clicked.connect(self.process_pdfs)
        process_layout.addWidget(self.process_btn)

        # Barra de progreso con etiqueta
        progress_layout = QVBoxLayout()
        self.progress_label = QLabel("Progreso:")
        self.progress_label.setVisible(False)
        progress_layout.addWidget(self.progress_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setMinimumHeight(20)
        progress_layout.addWidget(self.progress_bar)

        process_layout.addLayout(progress_layout)
        controls_layout.addWidget(process_group)

        # Sección de exportación
        export_group = QGroupBox("Exportación")
        export_group.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        export_layout = QVBoxLayout(export_group)

        # Botón de exportación
        self.export_btn = QPushButton("Exportar a Excel")
        self.export_btn.setMinimumHeight(35)
        self.export_btn.clicked.connect(self.export_results)
        self.export_btn.setEnabled(False)
        export_layout.addWidget(self.export_btn)

        controls_layout.addWidget(export_group)

        # Añadir contenedor de controles al splitter
        splitter.addWidget(controls_container)

        # =============== TABLA DE RESULTADOS ===============
        results_group = QGroupBox("Resultados de la extracción")
        results_group.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        results_layout = QVBoxLayout(results_group)

        # Tabla mejorada
        self.results_table = QTableWidget()
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.results_table.horizontalHeader().setMinimumSectionSize(100)

        results_layout.addWidget(self.results_table)

        # Añadir la tabla al splitter
        splitter.addWidget(results_group)

        # Configurar proporciones iniciales del splitter
        splitter.setSizes([250, 550])

        # Añadir splitter al layout principal
        main_layout.addWidget(splitter)

        self.setCentralWidget(central_widget)

    def select_pdfs(self):
        """Abre un diálogo para seleccionar archivos PDF"""
        file_dialog = QFileDialog()
        files, _ = file_dialog.getOpenFileNames(
            self,
            "Seleccionar archivos PDF",
            "",
            "Archivos PDF (*.pdf)"
        )

        if files:
            self.pdf_files = files
            file_count = len(files)

            # Actualizar etiquetas
            if file_count == 1:
                self.file_label.setText(f"Archivo seleccionado:")
                self.file_count_label.setText("1 PDF")
            else:
                self.file_label.setText(f"Archivos seleccionados:")
                self.file_count_label.setText(f"{file_count} PDFs")

            # Actualizar estado de botones
            self.process_btn.setEnabled(True)
            self.clear_btn.setEnabled(True)

            # Mostrar mensaje de estado
            self.statusBar.showMessage(f"{file_count} archivos PDF seleccionados correctamente")

    def clear_selection(self):
        """Limpia la selección de archivos actual"""
        self.pdf_files = []
        self.file_label.setText("No hay archivos seleccionados")
        self.file_count_label.setText("")
        self.process_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        self.statusBar.showMessage("Selección de archivos borrada")

    def process_pdfs(self):
        """Inicia el procesamiento de los PDFs seleccionados"""
        if not self.pdf_files:
            return

        # Configurar y mostrar barra de progreso
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.progress_label.setVisible(True)
        self.process_btn.setEnabled(False)
        self.select_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)

        # Mostrar mensaje en la barra de estado
        self.statusBar.showMessage("Procesando archivos PDF, por favor espere...")

        # Crear y configurar hilo de extracción
        self.extraction_thread = PDFExtractorThread(self.pdf_files)
        self.extraction_thread.progress_updated.connect(self.update_progress)
        self.extraction_thread.extraction_finished.connect(self.display_results)
        self.extraction_thread.error_occurred.connect(self.show_error)

        # Iniciar procesamiento
        self.extraction_thread.start()

    def update_progress(self, value):
        """Actualiza la barra de progreso"""
        self.progress_bar.setValue(value)
        if value < 100:
            self.statusBar.showMessage(f"Procesando... {value}% completado")
        else:
            self.statusBar.showMessage("Finalizando el procesamiento...")

    def display_results(self, dataframe):
        """Muestra los resultados en la tabla"""
        # Acortar nombres de columna para mejor visualización
        short_names = {}
        for col in dataframe.columns:
            if len(col) > 30:  # Si el nombre es muy largo
                short_name = col[:27] + "..."
                short_names[col] = short_name
            else:
                short_names[col] = col

        # Crear una copia del dataframe con nombres cortos para mostrar
        display_df = dataframe.copy()
        display_df.columns = [short_names[col] for col in dataframe.columns]

        # Configurar tabla con los resultados
        self.results_table.setRowCount(len(display_df))
        self.results_table.setColumnCount(len(display_df.columns))
        self.results_table.setHorizontalHeaderLabels(display_df.columns)

        # Configurar fuente del encabezado
        header_font = QFont("Arial", 9, QFont.Weight.Bold)
        self.results_table.horizontalHeader().setFont(header_font)

        # Llenar la tabla con datos
        for row in range(len(display_df)):
            for col in range(len(display_df.columns)):
                item = QTableWidgetItem(str(display_df.iloc[row, col]))
                self.results_table.setItem(row, col, item)

        # Guardar el dataframe original para exportación
        self.original_df = dataframe

        # Ajustar la tabla
        self.results_table.resizeColumnsToContents()

        # Habilitar/deshabilitar botones
        self.export_btn.setEnabled(True)
        self.process_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        self.clear_btn.setEnabled(True)

        # Ocultar elementos de progreso
        self.progress_bar.setVisible(False)
        self.progress_label.setVisible(False)

        # Actualizar barra de estado
        self.statusBar.showMessage(f"Procesamiento completado: {len(self.pdf_files)} archivos procesados")

        # Mostrar mensaje de éxito
        QMessageBox.information(self, "Proceso completado",
                                f"Se procesaron {len(self.pdf_files)} archivos PDF con éxito.\n"
                                f"Se extrajeron {len(dataframe.columns)} campos de datos.")

    def show_error(self, message):
        """Muestra un mensaje de error"""
        # Restaurar estado de la interfaz
        self.progress_bar.setVisible(False)
        self.progress_label.setVisible(False)
        self.process_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        self.clear_btn.setEnabled(True)

        # Actualizar barra de estado
        self.statusBar.showMessage("Error en el procesamiento")

        # Mostrar mensaje de error
        QMessageBox.critical(self, "Error", message)

    def apply_excel_styles(self, workbook, worksheet):
        """
        Aplica estilos al archivo Excel para mejorar su apariencia.

        Args:
            workbook: Libro de Excel (objeto openpyxl.Workbook)
            worksheet: Hoja de cálculo (objeto openpyxl.Worksheet)
        """
        # Definir estilos para los encabezados
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        # Definir estilos para las filas de datos
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(vertical='center', wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        # Colores alternos para las filas (celeste claro y blanco)
        light_blue_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

        # Aplicar estilo a los encabezados (primera fila)
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # Ajustar el alto de la primera fila
        worksheet.row_dimensions[1].height = 30

        # Aplicar estilos a las filas de datos con colores alternados
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border

                # Aplicar color alterno a las filas
                if row % 2 == 0:
                    cell.fill = light_blue_fill

        # Ajustar el ancho de las columnas para mejor visualización
        for col in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            # Encontrar la celda con el contenido más largo en esta columna
            for row in range(1, worksheet.max_row + 1):
                cell_value = str(worksheet.cell(row, col).value)
                max_length = max(max_length, len(cell_value))

            # Limitar el ancho máximo a 50 caracteres para evitar columnas demasiado anchas
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Congelar la primera fila
        worksheet.freeze_panes = "A2"

        # Aplicar autofilter para facilitar la navegación
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    def export_results(self):
        """Exporta los resultados a un archivo Excel con estilos mejorados"""
        if self.original_df is None or len(self.original_df) == 0:
            QMessageBox.warning(self, "Advertencia", "No hay datos para exportar.")
            return

        # Sugerir un nombre basado en la fecha
        import datetime
        default_name = f"datos_extraidos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar como Excel",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if file_path:
            try:
                # Mostrar progreso en la barra de estado
                self.statusBar.showMessage("Exportando datos a Excel...")

                # Crear un libro y hoja de trabajo de Excel usando openpyxl
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Datos Extraídos"

                # Añadir encabezados
                for col_idx, column_name in enumerate(self.original_df.columns, start=1):
                    worksheet.cell(row=1, column=col_idx).value = column_name

                # Añadir datos
                for row_idx, row in enumerate(self.original_df.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        # Evitar valores None
                        cell_value = "" if value is None else value
                        worksheet.cell(row=row_idx, column=col_idx).value = cell_value

                # Aplicar estilos
                self.apply_excel_styles(workbook, worksheet)

                # Guardar el archivo
                workbook.save(file_path)

                # Actualizar barra de estado
                self.statusBar.showMessage(f"Datos exportados exitosamente a {os.path.basename(file_path)}")

                # Mostrar mensaje de éxito
                QMessageBox.information(
                    self,
                    "Exportación exitosa",
                    f"Los datos fueron exportados correctamente a:\n{file_path}"
                )
            except Exception as e:
                self.statusBar.showMessage("Error en la exportación")
                QMessageBox.critical(
                    self,
                    "Error al exportar",
                    f"No se pudo exportar el archivo:\n{str(e)}"
                )

    def closeEvent(self, event):
        """Manejador para el cierre de la ventana"""
        # Detener el hilo si está en ejecución
        if hasattr(self, 'extraction_thread') and self.extraction_thread.isRunning():
            self.extraction_thread.stop()
            self.extraction_thread.wait()
        event.accept()